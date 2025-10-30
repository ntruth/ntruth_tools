"""TXT 文案转 Excel 工具。

该脚本既可以作为命令行程序使用，也可以在桌面环境下启动
一个简单的 GUI，帮助用户按照需求文档中的规则将 TXT 文案
整理到指定的 Excel 模板中。
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable, List, Optional

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:  # pragma: no cover - 在无图形环境下运行时跳过
    tk = None
    filedialog = messagebox = ttk = None  # type: ignore

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - 延迟到运行时报错
    load_workbook = None  # type: ignore


from templates import DEFAULT_TEMPLATE_FILENAME, ensure_default_template_file

CHINESE_COMMA = "，"
DEFAULT_TEMPLATE = Path(__file__).with_name("templates").joinpath(DEFAULT_TEMPLATE_FILENAME)


def extract_units(lines: Iterable[str]) -> List[str]:
    """根据需求文档的规则将原始行列表整理成文案单元列表。

    - 文案中的换行需要替换成中文逗号进行连接；
    - 遇到空行时视为一个文案单元结束；
    - 连续空行会被忽略，只作为一次分隔处理。
    """

    units: List[str] = []
    current: List[str] = []

    for raw_line in lines:
        # 使用 rstrip 去掉末尾的换行符，同时保留中间的空格和中文字符
        line = raw_line.rstrip("\n\r").replace("\ufeff", "")
        if line.strip() == "":
            if current:
                units.append(CHINESE_COMMA.join(current))
                current = []
            continue

        current.append(line.strip())

    if current:
        units.append(CHINESE_COMMA.join(current))

    return units


def convert_txt_to_excel(txt_path: Path, template_path: Path, output_path: Path, *, start_row: int = 1) -> int:
    """将 TXT 文案转换并写入 Excel。

    返回写入的文案单元数量。
    """

    if load_workbook is None:
        raise ModuleNotFoundError(
            "openpyxl 未安装，请先运行 `pip install openpyxl` 后再执行转换。"
        )

    if not txt_path.is_file():
        raise FileNotFoundError(f"未找到 TXT 文件：{txt_path}")

    if not template_path.is_file():
        raise FileNotFoundError(f"未找到 Excel 模板：{template_path}")

    with txt_path.open("r", encoding="utf-8") as fh:
        units = extract_units(fh.readlines())

    workbook = load_workbook(template_path)
    sheet = workbook.active

    if start_row < 1:
        raise ValueError("start_row 必须从 1 开始。")

    row_index = start_row
    for unit in units:
        sheet.cell(row=row_index, column=1).value = unit
        row_index += 1

    workbook.save(output_path)
    return len(units)


def run_cli(args: argparse.Namespace) -> None:
    txt_path = Path(args.txt)
    if args.template:
        template_path = Path(args.template)
    else:
        template_path = ensure_default_template_file(DEFAULT_TEMPLATE)
    output_path = Path(args.output)

    count = convert_txt_to_excel(txt_path, template_path, output_path, start_row=args.start_row)
    print(f"已写入 {count} 条文案到 {output_path}。")


def run_gui(default_template: Path = DEFAULT_TEMPLATE) -> None:  # pragma: no cover - GUI 难以自动化测试
    if tk is None:
        raise SystemExit("当前环境不支持 Tkinter 图形界面，请改用命令行参数运行。")

    default_template = ensure_default_template_file(default_template)

    root = tk.Tk()
    root.title("TXT 文案转 Excel 工具")
    root.resizable(False, False)

    mainframe = ttk.Frame(root, padding=16)
    mainframe.grid(column=0, row=0, sticky="nsew")

    for idx in range(3):
        mainframe.rowconfigure(idx, weight=1)
    mainframe.columnconfigure(1, weight=1)

    # 选择 TXT 文件
    ttk.Label(mainframe, text="TXT 文案文件：").grid(column=0, row=0, sticky="w", padx=(0, 8))
    txt_var = tk.StringVar()
    txt_entry = ttk.Entry(mainframe, width=48, textvariable=txt_var)
    txt_entry.grid(column=1, row=0, sticky="ew")

    def choose_txt() -> None:
        path = filedialog.askopenfilename(title="选择 TXT 文件", filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")])
        if path:
            txt_var.set(path)

    ttk.Button(mainframe, text="浏览...", command=choose_txt).grid(column=2, row=0, padx=(8, 0))

    # 选择模板
    ttk.Label(mainframe, text="Excel 模板：").grid(column=0, row=1, sticky="w", padx=(0, 8), pady=(8, 0))
    template_var = tk.StringVar(value=str(default_template))
    template_entry = ttk.Entry(mainframe, width=48, textvariable=template_var)
    template_entry.grid(column=1, row=1, sticky="ew", pady=(8, 0))

    def choose_template() -> None:
        path = filedialog.askopenfilename(title="选择 Excel 模板", filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")])
        if path:
            template_var.set(path)

    ttk.Button(mainframe, text="浏览...", command=choose_template).grid(column=2, row=1, padx=(8, 0), pady=(8, 0))

    # 输出路径
    ttk.Label(mainframe, text="输出文件：").grid(column=0, row=2, sticky="w", padx=(0, 8), pady=(8, 0))
    output_var = tk.StringVar()
    output_entry = ttk.Entry(mainframe, width=48, textvariable=output_var)
    output_entry.grid(column=1, row=2, sticky="ew", pady=(8, 0))

    def choose_output() -> None:
        path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if path:
            output_var.set(path)

    ttk.Button(mainframe, text="浏览...", command=choose_output).grid(column=2, row=2, padx=(8, 0), pady=(8, 0))

    status_var = tk.StringVar()
    status_label = ttk.Label(mainframe, textvariable=status_var, foreground="#666666")
    status_label.grid(column=0, row=3, columnspan=3, sticky="w", pady=(12, 0))

    def run_conversion() -> None:
        txt_value = txt_var.get().strip()
        output_value = output_var.get().strip()
        template_value = template_var.get().strip()

        if not txt_value:
            messagebox.showwarning("提示", "请先选择 TXT 文案文件。")
            return
        if not output_value:
            messagebox.showwarning("提示", "请选择 Excel 输出位置。")
            return

        txt_file = Path(txt_value)
        if not txt_file.is_file():
            messagebox.showerror("文件不存在", f"未找到 TXT 文件：{txt_file}")
            return

        template_file = Path(template_value) if template_value else ensure_default_template_file(default_template)
        if not template_file.is_file():
            messagebox.showerror("文件不存在", f"未找到 Excel 模板：{template_file}")
            return

        output_file = Path(output_value)

        try:
            count = convert_txt_to_excel(txt_file, template_file, output_file)
        except Exception as exc:  # pylint: disable=broad-except
            messagebox.showerror("转换失败", str(exc))
            status_var.set("")
            return

        messagebox.showinfo("转换完成", f"成功写入 {count} 条文案。")
        status_var.set(f"输出文件：{output_file}")

    action_frame = ttk.Frame(mainframe)
    action_frame.grid(column=0, row=4, columnspan=3, pady=(16, 0))
    ttk.Button(action_frame, text="开始转换", command=run_conversion).grid(column=0, row=0, padx=(0, 8))
    ttk.Button(action_frame, text="退出", command=root.destroy).grid(column=1, row=0)

    root.mainloop()


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="TXT 文案转 Excel 工具")
    parser.add_argument("--txt", help="需要转换的 TXT 文件路径")
    parser.add_argument("--template", help="Excel 模板路径，默认为仓库内置模板", default=None)
    parser.add_argument("--output", help="输出 Excel 文件路径")
    parser.add_argument("--start-row", type=int, default=1, help="在 Excel 中写入的起始行，默认为 1")
    parser.add_argument("--gui", action="store_true", help="强制启动图形界面")

    args = parser.parse_args(argv)

    if args.gui:
        return args

    if args.txt and args.output:
        return args

    provided_any = any(
        value is not None
        for value in (args.txt, args.output, args.template)
    ) or args.start_row != 1

    if provided_any:
        parser.error("命令行模式下必须同时提供 --txt 与 --output 参数，或使用 --gui 启动图形界面。")

    args.gui = True
    return args


def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv or sys.argv[1:])
    if args.gui:
        run_gui()
    else:
        run_cli(args)


if __name__ == "__main__":
    main()
